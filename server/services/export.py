import pandas as pd
import io
from models.student import Student
from database import db
from datetime import datetime

class ExportService:
    @staticmethod
    @staticmethod
    async def generate_excel(department: str = None, year: int = None, platform: str = None, contest_name: str = None, contest_date: str = None):
        from .platforms.codeforces import CodeforcesService  # Import here to avoid circular dep if any
        
        query = {}
        if department and department != "All":
            query["department"] = department
        if year and year != "All":
            query["year"] = int(year)

        students = list(db.get_db()["students"].find(query))
        
        # Sort by Reg No (case-insensitive alphanumeric sort)
        try:
             students.sort(key=lambda s: s.get("reg_no", "").lower())
        except:
             pass
        
        output = io.BytesIO()
        
        # --- Contest Report Mode ---
        if platform and contest_name:
            data = []
            selected_platform = platform.lower()
            
            # --- Codeforces Live Fetch Logic ---
            cf_standings = None
            cf_problems = None
            if selected_platform == "codeforces" and contest_name.isdigit():
                 # Valid Contest ID, try fetching live data
                 handles = [s.get("handles", {}).get("codeforces") for s in students if s.get("handles", {}).get("codeforces")]
                 handles = [h for h in handles if h] # Filter None
                 
                 cf_rows, cf_probs = await CodeforcesService.get_contest_standings(contest_name, handles)
                 if cf_rows:
                     # Map by handle (lowercase)
                     cf_standings = {}
                     for r in cf_rows:
                         for m in r["party"]["members"]:
                             h_low = m["handle"].lower()
                             cf_standings[h_low] = r
                     cf_problems = cf_probs

            for idx, s in enumerate(students, 1):
                base_info = {
                    "S. No": idx,
                    "Register Number": s.get("reg_no", "Unknown"),
                    "Name of the Student": s.get("name", "Unknown"),
                }
                stats = s.get("stats", {}).get(selected_platform, {})
                
                row = {}
                if selected_platform == "leetcode":
                    # Try to find specific contest data
                    history = stats.get("history", [])
                    contest_stats = None
                    
                    if contest_name:
                        # Fuzzy match or exact match contest name
                        # Standardize: lowercase, remove spaces, hyphens, underscores
                        def clean_str(s):
                            return str(s).lower().replace(" ", "").replace("-", "").replace("_", "").replace(".", "")
                            
                        clean_name = clean_str(contest_name)
                        for h in history:
                            c_title = clean_str(h.get("contest", {}).get("title", ""))
                            if clean_name in c_title:
                                contest_stats = h
                                break
                    
                    # If found, use contest specific stats
                    solved_count = contest_stats.get("problemsSolved", 0) if contest_stats else 0
                    contest_rating = contest_stats.get("rating", "Absent") if contest_stats else "Absent"
                    
                    # Note: LeetCode API doesn't give Easy/Medium/Hard breakdown per contest in this history, just total solved (0-4).
                    # We will put total solved in 'Total' and 0 in others to avoid confusion, or try to distribute if we knew better.
                    # For now, we put the count in Total.
                    
                    # Calculate Top % for Specific Contest
                    top_percent_str = "N/A"
                    if contest_stats:
                        rank = contest_stats.get("ranking", 0)
                        total_participants = contest_stats.get("totalParticipants", 0)
                        if total_participants > 0 and rank > 0:
                            top_percent = (rank / total_participants) * 100
                            top_percent_str = f"{top_percent:.2f}%"

                    row = {
                        **base_info,
                        "Leet Code Easy": "-", # Not available in simple history
                        "Leet Code Medium": "-",
                        "Leet code Hard": "-",
                        "Total": solved_count if contest_stats else "Absent",
                        "Contest count": stats.get("attended", 0), # Overall attended
                        "Contest Rating": contest_rating,
                        "Global Rank": stats.get("global_rank", "N/A"),
                        "Top %": top_percent_str
                    }
                elif selected_platform == "codeforces":
                    # Check if we have live data
                    unique_cf_row = None
                    if cf_standings:
                        handle = s.get("handles", {}).get("codeforces", "").lower()
                        unique_cf_row = cf_standings.get(handle)
                    
                    if unique_cf_row:
                        # Use Live Data
                        points = unique_cf_row.get("points", 0)
                        rank = unique_cf_row.get("rank", 0)
                        penalty = unique_cf_row.get("penalty", 0)
                        
                        # Calculate solved count from problemResults
                        solved_cnt = 0
                        problem_results = unique_cf_row.get("problemResults", [])
                        
                        row = {
                            **base_info,
                            "Rank": rank,
                            "Points": points,
                            "Penalty": penalty
                        }
                        
                        # Add dynamic problem columns
                        if cf_problems:
                            for i, p in enumerate(cf_problems):
                                p_idx = p.get("index", str(i))
                                res = problem_results[i] if i < len(problem_results) else {}
                                points_got = res.get("points", 0)
                                if points_got > 0:
                                    row[p_idx] = points_got
                                    solved_cnt += 1
                                else:
                                    row[p_idx] = 0
                        
                        row["Total Solved"] = solved_cnt
                        
                    else:
                        # Fallback to stored history
                        history = stats.get("history", [])
                        contest_stats = None
                        if contest_name:
                                def clean_str(s):
                                    return str(s).lower().replace(" ", "").replace("-", "").replace("_", "").replace(".", "")

                                clean_name = clean_str(contest_name)
                                for h in history:
                                    # Check Contest Name
                                    c_title = clean_str(h.get("contestName", ""))
                                    # Check Contest ID
                                    c_id = str(h.get("contestId", ""))
                                    
                                    if clean_name in c_title or clean_name == c_id:
                                        contest_stats = h
                                        break

                        row = {
                            **base_info,
                            "Contest Rank": contest_stats.get("rank", "Absent") if contest_stats else "Absent",
                            "Rating After": contest_stats.get("newRating", "Absent") if contest_stats else "Absent",
                            "Rating Change": (contest_stats.get("newRating", 0) - contest_stats.get("oldRating", 0)) if contest_stats else "-",
                            "Current Rating": stats.get("rating", 0),
                            "Total Solved": stats.get("solved", 0)
                        }
                        
                elif selected_platform == "codechef":
                     history = stats.get("history", [])
                     contest_stats = None
                     if contest_name:
                         clean_name = contest_name.lower().replace(" ", "")
                         for h in history:
                             # Codechef history has 'name' (e.g. Starters 100) and 'code' (e.g. START100)
                             c_title = h.get("name", "").lower().replace(" ", "")
                             c_code = h.get("code", "").lower()
                             if clean_name in c_title or clean_name in c_code:
                                 contest_stats = h
                                 break
                    
                     row = {
                        **base_info,
                        "Contest Rank": contest_stats.get("rank", "Absent") if contest_stats else "Absent",
                        "Contest Rating": contest_stats.get("rating", "Absent") if contest_stats else "Absent",
                        "Current Rating": stats.get("rating", 0),
                        "Global Rank": stats.get("global_rank", "N/A"),
                        "Total Solved": stats.get("solved", 0)
                    }
                else:
                    # Default/HackerRank
                     row = {
                        **base_info,
                        "Badges": stats.get("badges", 0),
                        "Solved": stats.get("solved", 0)
                    }
                
                data.append(row)

            if not data:
                # Handle empty data case with default columns to avoid crash
                if selected_platform == "leetcode":
                    cols = ["S. No", "Register Number", "Name of the Student", "Leet Code Easy", "Leet Code Medium", "Leet code Hard", "Total", "Contest count", "Contest Rating", "Global Rank", "Top %"]
                elif selected_platform == "codeforces":
                    cols = ["S. No", "Register Number", "Name of the Student", "Contest Rank", "Rating After", "Rating Change", "Current Rating", "Total Solved"]
                elif selected_platform == "codechef":
                    cols = ["S. No", "Register Number", "Name of the Student", "Contest Rank", "Contest Rating", "Current Rating", "Global Rank", "Total Solved"]
                else:
                    cols = ["S. No", "Register Number", "Name of the Student", "Badges", "Solved"]
                df = pd.DataFrame(columns=cols)
            else:
                df = pd.DataFrame(data)
            
            # Use openpyxl for custom header
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                # Write data starting from row 3 (leaving room for header)
                df.to_excel(writer, index=False, sheet_name='Contest Data', startrow=2)
                
                workbook = writer.book
                worksheet = writer.sheets['Contest Data']
                
                # Add Big Date Header
                from openpyxl.styles import Font, Alignment
                
                # Merge first row
                worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(df.columns))
                cell = worksheet.cell(row=1, column=1)
                cell.value = contest_date if contest_date else datetime.now().strftime("%d-%m-%Y")
                cell.font = Font(size=24, bold=True)
                cell.alignment = Alignment(horizontal='center', vertical='center')
                
                # Adjust column widths
                from openpyxl.utils import get_column_letter
                for i, col in enumerate(worksheet.columns, 1):
                    max_length = 0
                    column_letter = get_column_letter(i)
                    for cell in col:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = (max_length + 2)
                    worksheet.column_dimensions[column_letter].width = adjusted_width

            output.seek(0)
            return output

        # --- Normal Export Mode (Multi-sheet) ---
        
        # Prepare lists for each sheet
        lc_data = []
        cf_data = []
        cc_data = []
        hr_data = []
        
        for idx, s in enumerate(students, 1):
            base_info = {
                "S.No": idx,
                "Reg No": s.get("reg_no", "Unknown"),
                "Name": s.get("name", "Unknown"),
                "Department": s.get("department", "Unknown")
            }
            stats = s.get("stats", {})
            
            # LeetCode Row
            lc = stats.get("leetcode", {})
            lc_row = {
                **base_info,
                "LeetCode Easy": lc.get("easy", 0),
                "LeetCode Medium": lc.get("medium", 0),
                "LeetCode Hard": lc.get("hard", 0),
                "Total Solved": lc.get("total_solved", 0),
                "Contest Count": lc.get("attended", 0),
                "Contest Rating": lc.get("rating", "N/A"),
                "Global Rank": lc.get("global_rank", "N/A"),
                "Top %": f"{lc.get('top_percentage', 0)}%" if lc.get('top_percentage') else "N/A"
            }
            lc_data.append(lc_row)
            
            # CodeChef Row
            cc = stats.get("codechef", {})
            cc_row = {
                **base_info,
                "Rating": cc.get("rating", 0),
                "Global Rank": cc.get("global_rank", "N/A"),
                "Stars": cc.get("stars", 0),
                "Solved": cc.get("solved", 0),
                "Contests": cc.get("contests", 0)
            }
            cc_data.append(cc_row)
            
            # Codeforces Row
            cf = stats.get("codeforces", {})
            cf_row = {
                **base_info,
                "Rating": cf.get("rating", 0),
                "Max Rating": cf.get("max_rating", 0),
                "Rank": cf.get("rank", "Unrated"),
                "Total Solved": cf.get("solved", 0),
                "Contests": cf.get("contests", 0)
            }
            cf_data.append(cf_row)
            
            # HackerRank Row
            hr = stats.get("hackerrank", {})
            hr_row = {
                **base_info,
                "Badges": hr.get("badges", 0),
                "Solved": hr.get("solved", 0)
            }
            hr_data.append(hr_row)
            
        # Create DataFrames
        df_lc = pd.DataFrame(lc_data)
        df_cc = pd.DataFrame(cc_data)
        df_cf = pd.DataFrame(cf_data)
        df_hr = pd.DataFrame(hr_data)
        
        # Write to Excel
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            if not df_lc.empty:
                df_lc.to_excel(writer, index=False, sheet_name='LeetCode')
            else:
                 pd.DataFrame(columns=["S.No", "Name", "Department"]).to_excel(writer, index=False, sheet_name='LeetCode')
                 
            if not df_cc.empty:
                df_cc.to_excel(writer, index=False, sheet_name='CodeChef')
            else:
                pd.DataFrame(columns=["S.No", "Name", "Department"]).to_excel(writer, index=False, sheet_name='CodeChef')

            if not df_cf.empty:
                df_cf.to_excel(writer, index=False, sheet_name='Codeforces')
            else:
                 pd.DataFrame(columns=["S.No", "Name", "Department"]).to_excel(writer, index=False, sheet_name='Codeforces')

            if not df_hr.empty:
                df_hr.to_excel(writer, index=False, sheet_name='HackerRank')
            else:
                 pd.DataFrame(columns=["S.No", "Name", "Department"]).to_excel(writer, index=False, sheet_name='HackerRank')
            
        output.seek(0)
        return output
